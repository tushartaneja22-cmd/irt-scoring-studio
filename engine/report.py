"""Write calibration results to Excel / CSV and compute validation metrics."""
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def metrics(est, gold):
    est = np.asarray(est, float); gold = np.asarray(gold, float)
    m = ~(np.isnan(est) | np.isnan(gold))
    est, gold = est[m], gold[m]
    if len(est) < 2:
        return dict(r=float('nan'), rmse=float('nan'), mae=float('nan'), n=len(est))
    return dict(r=float(np.corrcoef(est, gold)[0, 1]),
                rmse=float(np.sqrt(np.mean((est - gold) ** 2))),
                mae=float(np.mean(np.abs(est - gold))), n=len(est))


def write_results_xlsx(results, path, gold=None, mock_id=None):
    """results: dict subject -> {'items':[...]}. Writes one sheet per subject
    plus a values-only sheet matching the gold layout (a,b,c columns)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_font = Font(bold=True, color='FFFFFF', name='Arial')
    hdr_fill = PatternFill('solid', fgColor='305496')
    flag_fill = PatternFill('solid', fgColor='FCE4D6')
    for subj in ('rw', 'math'):
        if subj not in results:
            continue
        items = results[subj]['items']
        ws = wb.create_sheet(f"{subj}_detail")
        corrected = bool(items) and 'a_uncorrected' in items[0]
        cols = ['qid', 'section', 'itemtype', 'n', 'p', 'a', 'b', 'c']
        if corrected:
            cols += ['a_uncorrected', 'b_uncorrected', 'c_uncorrected',
                     'a_gold', 'b_gold', 'c_gold', 'corrected']
        cols += ['a_raw', 'b_raw', 'flags']
        for ci, cname in enumerate(cols, 1):
            cell = ws.cell(1, ci, cname); cell.font = hdr_font; cell.fill = hdr_fill
        for ri, rec in enumerate(items, 2):
            for ci, cname in enumerate(cols, 1):
                ws.cell(ri, ci, rec.get(cname))
            if rec.get('flags'):
                for ci in range(1, len(cols) + 1):
                    ws.cell(ri, ci).fill = flag_fill
        ws.freeze_panes = 'A2'
        from openpyxl.utils import get_column_letter
        for ci in range(1, len(cols) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 11
        # values-only sheet (gold format: a,b,c)
        wv = wb.create_sheet(f"{subj}_abc")
        for ci, nm in enumerate(['a', 'b', 'c'], 1):
            c = wv.cell(1, ci, nm); c.font = hdr_font; c.fill = hdr_fill
        for ri, rec in enumerate(items, 2):
            wv.cell(ri, 1, rec['a']); wv.cell(ri, 2, rec['b']); wv.cell(ri, 3, rec['c'])
    wb.save(path)
    return path


def validation_table(results, gold, mock_id):
    """Return list of rows: subject, param, r, rmse, mae, n (aligned only)."""
    rows = []
    for subj in ('rw', 'math'):
        key = (mock_id, subj)
        if subj not in results or key not in gold:
            continue
        items = results[subj]['items']
        g = gold[key]
        if g.shape[0] != len(items):
            rows.append(dict(subject=subj, param='-', note=f'misaligned est {len(items)} vs gold {g.shape[0]}'))
            continue
        for pi, pname in enumerate(['a', 'b', 'c']):
            est = [it[pname] for it in items]
            m = metrics(est, g[:, pi])
            rows.append(dict(subject=subj, param=pname, **m))
    return rows
