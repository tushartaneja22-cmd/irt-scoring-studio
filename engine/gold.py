"""Read reference (gold-standard) a,b,c values and provide comparison metrics."""
import numpy as np
import openpyxl

MOCKS = ['115', '116', '117', '118', '119', '120']


def load_gold(path='A, B, C Values.xlsx'):
    wb = openpyxl.load_workbook(path, data_only=True)
    gold = {}
    for name in wb.sheetnames:
        num, subj = name.split()  # "115 English"
        ws = wb[name]
        rows = []
        for r in range(2, ws.max_row + 1):
            vals = [ws.cell(r, c).value for c in range(1, 4)]
            if vals[0] is None:
                continue
            rows.append([float(v) for v in vals])
        arr = np.array(rows)  # (J,3) a,b,c
        key = (num, 'rw' if subj.lower().startswith('eng') else 'math')
        gold[key] = arr
    return gold


def compare(est_a, est_b, est_c, gold_arr):
    ga, gb, gc = gold_arr[:, 0], gold_arr[:, 1], gold_arr[:, 2]
    def stats(e, g):
        r = np.corrcoef(e, g)[0, 1]
        rmse = np.sqrt(np.mean((e - g) ** 2))
        mae = np.mean(np.abs(e - g))
        return r, rmse, mae
    return {
        'a': stats(est_a, ga),
        'b': stats(est_b, gb),
        'c': stats(est_c, gc),
    }


def fmt(name, m):
    out = [name]
    for k in ('a', 'b', 'c'):
        r, rmse, mae = m[k]
        out.append(f"{k}: r={r:.3f} rmse={rmse:.3f} mae={mae:.3f}")
    return "  |  ".join(out)
