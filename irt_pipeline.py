"""
IRT Scoring Pipeline  -  Calibrate (3PL) + Score (WLE)  -  reusable
==================================================================

Give it a response file (CSV or Excel), get back ONE Excel workbook with two
tabs:
    * "Item Parameters" - a, b, c (+ SE, n, confidence) for every question
    * "Student Scores"  - each student's ability (theta) and scaled score per
                          section, plus a total

    python irt_pipeline.py responses.csv
    python irt_pipeline.py workbook.xlsx --responses-sheet "Student Responses"
    python irt_pipeline.py responses.csv --out results.xlsx
    python irt_pipeline.py responses.csv --scale "R&W:670:60,Math:710:70"
    python irt_pipeline.py responses.csv --scale-json assessment.json

INPUT FORMAT (works for any test, not just this one)
----------------------------------------------------
A table with one row per student. Columns:
  * an ID column and (optionally) a Name column at the front, then
  * one column per question, each cell = 1 (correct), 0 (wrong),
    or "-"/blank (not administered - e.g. an adaptive form the student
    didn't take). Any non-0/1 value is treated as "not administered".

By default question columns are auto-detected: first it tries the pattern
S<sec>_<id>_...  (e.g. S1_28575_(STANDARD)(reading_and_writing)); if that
matches nothing, it falls back to "every column whose values are only 0/1/-".
Each question's SECTION is inferred from its column name (math vs
reading/writing/verbal); override the keyword map with --subject-keywords.

METHOD
------
  Stage 1  Calibrate 3PL item parameters by Marginal Maximum Likelihood
           (Bock-Aitkin EM). D=1.702 (normal-ogive metric); mild priors on
           a (log-normal), b (normal), c (Beta, mode 0.25 = 1/4 MCQ guess).
  Stage 2  Estimate each student's theta per section with WLE (Warm's
           Weighted Likelihood - bias-corrected, defined even for perfect
           scores), then scale:  score = xbar + sigma * theta,  rounded to
           the nearest 10 and clamped to [200, 800].

Nothing is required from you but the responses; the ability scale is fixed
internally to mean 0 / SD 1. The section xbar/sigma only convert theta to a
reported score and default to Digital-SAT values (R&W 670/60, Math 710/70).
"""

import argparse
import csv
import json
import os
import re

import numpy as np
from scipy.optimize import minimize, brentq
from scipy.special import logsumexp
from scipy.stats import norm, beta as beta_dist, lognorm
from openpyxl import load_workbook, Workbook

# ----------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------
N_QUAD = 41
THETA_MIN, THETA_MAX = -4.0, 4.0
D = 1.702                       # logistic scaling constant (normal-ogive metric)

N_HIGH, N_MODERATE = 200, 30    # confidence tiers by sample size

A_BOUNDS, B_BOUNDS, C_BOUNDS = (0.05, 4.0), (-4.0, 4.0), (1e-3, 0.5)
A_PRIOR_SD, B_PRIOR_SD = 0.5, 4.0
C_PRIOR_A, C_PRIOR_B = 5, 13    # Beta(5,13), mode 0.25

MAX_EM_ITER, LL_TOL, _EPS = 500, 1e-4, 1e-6

DEFAULT_SCALE = {"R&W": (670.0, 60.0), "Math": (710.0, 70.0), "Other": (500.0, 100.0)}
SCORE_MIN, SCORE_MAX, SCORE_STEP = 200, 800, 10

DEFAULT_SUBJECT_KEYWORDS = {
    "Math": ["math"],
    "R&W": ["reading", "writing", "verbal", "english", "r&w", "rw"],
}

ITEM_COL_PATTERN = re.compile(r"^S\d+_\d+_")
_QNUM = re.compile(r"_(\d+)_")

quad = np.linspace(THETA_MIN, THETA_MAX, N_QUAD)
_pw = norm.pdf(quad, 0.0, 1.0); _pw /= _pw.sum()
_log_pw = np.log(_pw)[None, :]


# ----------------------------------------------------------------------
# 3PL core (+ derivatives for scoring)
# ----------------------------------------------------------------------
def p_3pl(theta, a, b, c):
    return c + (1.0 - c) / (1.0 + np.exp(-D * a * (theta - b)))


def p_and_derivs(theta, a, b, c):
    psi = 1.0 / (1.0 + np.exp(-D * a * (theta - b)))
    P = c + (1.0 - c) * psi
    dP = (1.0 - c) * D * a * psi * (1.0 - psi)
    d2P = (1.0 - c) * (D * a) ** 2 * psi * (1.0 - psi) * (1.0 - 2.0 * psi)
    return P, dP, d2P


# ----------------------------------------------------------------------
# Stage 1: calibration
# ----------------------------------------------------------------------
def _log_prior(a, b, c):
    return (lognorm.logpdf(a, s=A_PRIOR_SD, scale=1.0)
            + norm.logpdf(b, 0.0, B_PRIOR_SD)
            + beta_dist.logpdf(np.clip(c, _EPS, 1 - _EPS), C_PRIOR_A, C_PRIOR_B))


def _neg_log_post(params, n_k, r_k):
    a, b, c = params
    p = np.clip(p_3pl(quad, a, b, c), _EPS, 1 - _EPS)
    ll = np.sum(r_k * np.log(p) + (n_k - r_k) * np.log(1 - p))
    return -(ll + _log_prior(a, b, c))


def _se_b(params, n_k, r_k):
    x = np.asarray(params, float); h = 1e-4; H = np.zeros((3, 3))
    for i in range(3):
        for j in range(3):
            d = np.zeros(3)
            def f(si, sj):
                y = x.copy(); y[i] += si * h; y[j] += sj * h
                return _neg_log_post(y, n_k, r_k)
            H[i, j] = (f(1, 1) - f(1, -1) - f(-1, 1) + f(-1, -1)) / (4 * h * h)
    try:
        v = np.linalg.inv(H)[1, 1]
        return float(np.sqrt(v)) if v > 0 else float("nan")
    except np.linalg.LinAlgError:
        return float("nan")


def _posterior(resp, mask, a, b, c):
    n = resp.shape[1]
    log_lik = np.zeros((n, N_QUAD))
    for k, tk in enumerate(quad):
        p = np.clip(p_3pl(tk, a, b, c), _EPS, 1 - _EPS)
        log_lik[:, k] = (mask * (resp * np.log(p)[:, None] + (1 - resp) * np.log(1 - p)[:, None])).sum(0)
    w = log_lik + _log_pw
    mll = float(logsumexp(w, axis=1).sum())
    lp = w - w.max(1, keepdims=True)
    post = np.exp(lp); post /= post.sum(1, keepdims=True)
    return post, mll


def calibrate(resp, mask, verbose=True):
    n_items = resp.shape[0]
    n_adm = mask.sum(1)
    a, b, c = np.ones(n_items), np.zeros(n_items), np.full(n_items, 0.25)
    converged, it, prev = False, 0, -np.inf
    for it in range(1, MAX_EM_ITER + 1):
        post, mll = _posterior(resp, mask, a, b, c)
        n_jk, r_jk = mask @ post, (mask * resp) @ post
        chg = 0.0; na, nb, nc = a.copy(), b.copy(), c.copy()
        for j in range(n_items):
            if n_adm[j] == 0:
                continue
            r = minimize(_neg_log_post, [a[j], b[j], c[j]], args=(n_jk[j], r_jk[j]),
                         method="L-BFGS-B", bounds=[A_BOUNDS, B_BOUNDS, C_BOUNDS])
            chg = max(chg, *np.abs(r.x - [a[j], b[j], c[j]]))
            na[j], nb[j], nc[j] = r.x
        a, b, c = na, nb, nc
        if verbose:
            print(f"  EM iter {it:3d}: logL={mll:.3f}  max change={chg:.5f}")
        if it > 1 and abs(mll - prev) < LL_TOL:
            converged = True; break
        prev = mll
    post, _ = _posterior(resp, mask, a, b, c)
    n_jk, r_jk = mask @ post, (mask * resp) @ post
    se = np.array([_se_b([a[j], b[j], c[j]], n_jk[j], r_jk[j]) if n_adm[j] else np.nan
                   for j in range(n_items)])
    return a, b, c, se, n_adm, it, converged


# ----------------------------------------------------------------------
# Stage 2: WLE scoring
# ----------------------------------------------------------------------
def wle_theta(a, b, c, u):
    """Warm's weighted-likelihood theta for one student's responses in one
    section. Falls back to the search-range edge for all-right/all-wrong."""
    a, b, c, u = map(np.asarray, (a, b, c, u))
    def S(t):
        P, dP, _ = p_and_derivs(t, a, b, c); P = np.clip(P, 1e-9, 1 - 1e-9)
        return np.sum(dP * (u - P) / (P * (1 - P)))
    def info(t):
        P, dP, _ = p_and_derivs(t, a, b, c); P = np.clip(P, 1e-9, 1 - 1e-9)
        return np.sum(dP ** 2 / (P * (1 - P)))
    def Jt(t):
        P, dP, d2P = p_and_derivs(t, a, b, c); P = np.clip(P, 1e-9, 1 - 1e-9)
        return np.sum(dP * d2P / (P * (1 - P)))
    def g(t):
        I = info(t)
        return S(t) + (Jt(t) / (2 * I) if I > 1e-12 else 0.0)
    xs = np.linspace(-6, 6, 481)
    gv = np.array([g(t) for t in xs])
    for k in range(len(xs) - 1):
        if gv[k] == 0:
            return float(xs[k])
        if gv[k] * gv[k + 1] < 0:
            return float(brentq(g, xs[k], xs[k + 1]))
    return float(xs[int(np.argmin(np.abs(gv)))])  # monotonic -> edge


def score_students(resp, mask, item_subject, a, b, c, subjects):
    """Return {subject: (theta[n], n_items[n])} using WLE per student."""
    n_students = resp.shape[1]
    out = {}
    for s in subjects:
        idx = np.where(item_subject == s)[0]
        if len(idx) == 0:
            continue
        theta = np.zeros(n_students); nit = np.zeros(n_students)
        for i in range(n_students):
            m = mask[idx, i].astype(bool)
            nit[i] = m.sum()
            if not m.any():
                theta[i] = np.nan; continue
            theta[i] = wle_theta(a[idx][m], b[idx][m], c[idx][m], resp[idx, i][m])
        out[s] = (theta, nit)
    return out


def to_scaled(theta, xbar, sigma):
    raw = np.clip(xbar + sigma * theta, SCORE_MIN, SCORE_MAX)
    return (np.round(raw / SCORE_STEP) * SCORE_STEP).astype(int)


# ----------------------------------------------------------------------
# Reading any response table (CSV or Excel)
# ----------------------------------------------------------------------
def read_table(path, sheet):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".txt", ".tsv"):
        delim = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f, delimiter=delim))
        return rows[0], rows[1:]
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    body = [[ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            for r in range(2, ws.max_row + 1)]
    return header, body


def _resp_val(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return int(v) if v in (0, 1) else None
    s = str(v).strip()
    if s in ("0", "1"):
        return int(s)
    if s in ("0.0", "1.0"):
        return int(float(s))
    return None


def detect_item_columns(header, body):
    cols = [c for c, name in enumerate(header) if name and ITEM_COL_PATTERN.match(str(name))]
    if cols:
        return cols
    # fallback: columns whose non-empty values are only 0/1/-/blank and have >=1 response
    cols = []
    for c in range(len(header)):
        vals = [row[c] if c < len(row) else None for row in body]
        parsed = [_resp_val(v) for v in vals]
        nonblank = [v for v in vals if v not in (None, "") and str(v).strip() != ""]
        has01 = any(p is not None for p in parsed)
        allok = all((_resp_val(v) is not None) or str(v).strip() in ("-", "") for v in nonblank)
        if has01 and allok:
            cols.append(c)
    return cols


def subject_of(name, keyword_map):
    n = str(name).lower()
    for subj, kws in keyword_map.items():
        if any(k in n for k in kws):
            return subj
    return "Other"


# ----------------------------------------------------------------------
# Writing the two-tab workbook
# ----------------------------------------------------------------------
def _conf(n):
    return "High" if n >= N_HIGH else "Moderate" if n >= N_MODERATE else "Low"


def write_workbook(path, item_ids, item_subject, a, b, c, se, n_adm,
                   ids, names, scored, scale, subjects):
    wb = Workbook()
    ws1 = wb.active; ws1.title = "Item Parameters"
    ws1.append(["Question ID", "Subject", "a (Discrimination)", "b (Difficulty)",
                "c (Guessing)", "SE (b)", "n students", "Confidence"])
    for j, qid in enumerate(item_ids):
        n = int(n_adm[j])
        if n == 0:
            ws1.append([qid, item_subject[j], None, None, None, None, 0, "None"]); continue
        ws1.append([qid, item_subject[j], round(float(a[j]), 3), round(float(b[j]), 3),
                    round(float(c[j]), 3),
                    round(float(se[j]), 3) if np.isfinite(se[j]) else "n/a", n, _conf(n)])

    ws2 = wb.create_sheet("Student Scores")
    head = ["Id", "Name"]
    for s in subjects:
        head += [f"{s} items", f"{s} theta", f"{s} score"]
    head += ["Total score"]
    ws2.append(head)
    for i in range(len(ids)):
        row = [ids[i], names[i]]; total = 0
        for s in subjects:
            theta, nit = scored[s]
            if np.isnan(theta[i]):
                row += [int(nit[i]), None, None]; continue
            xb, sg = scale.get(s, DEFAULT_SCALE["Other"])
            sc = int(to_scaled(np.array([theta[i]]), xb, sg)[0])
            row += [int(nit[i]), round(float(theta[i]), 3), sc]; total += sc
        row += [total]
        ws2.append(row)
    wb.save(path)


# ----------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------
def parse_scale(arg):
    scale = dict(DEFAULT_SCALE)
    if arg:
        for part in arg.split(","):
            name, xb, sg = part.split(":")
            scale[name.strip()] = (float(xb), float(sg))
    return scale


def scale_from_json(path, scale):
    d = json.load(open(path, encoding="utf-8")).get("assessment", {})
    if "english_xbar" in d:
        scale["R&W"] = (float(d["english_xbar"]), float(d["english_sigma"]))
    if "math_xbar" in d:
        scale["Math"] = (float(d["math_xbar"]), float(d["math_sigma"]))
    return scale


def main():
    ap = argparse.ArgumentParser(description="IRT pipeline: 3PL calibration + WLE scoring -> two-tab workbook.")
    ap.add_argument("input", help="responses file (.csv, .tsv, or .xlsx)")
    ap.add_argument("--responses-sheet", default="Student Responses", help="sheet name if input is Excel")
    ap.add_argument("--out", help="output .xlsx (default: <input>_results.xlsx)")
    ap.add_argument("--id-col", type=int, default=0, help="0-based index of the ID column")
    ap.add_argument("--name-col", type=int, default=1, help="0-based index of the Name column (-1 if none)")
    ap.add_argument("--scale", help='per-section "Name:xbar:sigma,..." e.g. "R&W:670:60,Math:710:70"')
    ap.add_argument("--scale-json", help="assessment JSON providing english/math xbar/sigma")
    ap.add_argument("--subject-keywords", help='override, e.g. "Math=math;R&W=reading,writing,verbal"')
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    scale = parse_scale(args.scale)
    if args.scale_json:
        scale = scale_from_json(args.scale_json, scale)
    kw = dict(DEFAULT_SUBJECT_KEYWORDS)
    if args.subject_keywords:
        kw = {}
        for grp in args.subject_keywords.split(";"):
            name, words = grp.split("=")
            kw[name.strip()] = [w.strip().lower() for w in words.split(",")]

    out = args.out or f"{os.path.splitext(args.input)[0]}_results.xlsx"

    print(f"Reading {args.input} ...")
    header, body = read_table(args.input, args.responses_sheet)
    item_cols = detect_item_columns(header, body)
    if not item_cols:
        raise SystemExit("No question columns detected. Check the file or --item pattern.")
    item_ids = [str(header[c]) for c in item_cols]
    item_subject = np.array([subject_of(header[c], kw) for c in item_cols])

    n_items, n_students = len(item_cols), len(body)
    resp = np.zeros((n_items, n_students)); mask = np.zeros((n_items, n_students))
    for j, c in enumerate(item_cols):
        for i, row in enumerate(body):
            v = _resp_val(row[c] if c < len(row) else None)
            if v is not None:
                resp[j, i] = v; mask[j, i] = 1
    ids = [row[args.id_col] if args.id_col < len(row) else None for row in body]
    names = ([row[args.name_col] if 0 <= args.name_col < len(row) else None for row in body]
             if args.name_col >= 0 else [None] * n_students)

    subjects = [s for s in list(dict.fromkeys(item_subject)) ]  # preserve order of appearance
    counts = {s: int((item_subject == s).sum()) for s in subjects}
    print(f"  {n_items} questions x {n_students} students | sections: {counts}")
    missing_scale = [s for s in subjects if s not in scale]
    if missing_scale:
        print(f"  NOTE: no scale given for {missing_scale}; using default {DEFAULT_SCALE['Other']}")

    print("STAGE 1 - calibrating 3PL item parameters (Bock-Aitkin EM) ...")
    a, b, c, se, n_adm, it, conv = calibrate(resp, mask, verbose=not args.quiet)
    print(f"  {'converged' if conv else 'STOPPED at cap'} after {it} iterations")

    print("STAGE 2 - scoring students (WLE theta per section) ...")
    scored = score_students(resp, mask, item_subject, a, b, c, subjects)
    for s in subjects:
        theta = scored[s][0]; valid = theta[~np.isnan(theta)]
        xb, sg = scale.get(s, DEFAULT_SCALE["Other"])
        sc = to_scaled(valid, xb, sg)
        print(f"  {s:5s}: scale {xb:.0f}+{sg:.0f}*theta | mean score {sc.mean():.0f}, "
              f"range {sc.min()}-{sc.max()}")

    write_workbook(out, item_ids, item_subject, a, b, c, se, n_adm,
                   ids, names, scored, scale, subjects)
    print(f"Done -> {out}  (tabs: 'Item Parameters', 'Student Scores')")


if __name__ == "__main__":
    main()
